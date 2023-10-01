import json
from datetime import datetime
from django.db.models import Sum
from django.dispatch import receiver
from django.http import JsonResponse, HttpResponse
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.workbook import Workbook

# Импортируем модели и сигналы из приложений "robots" и "orders"
from .models import Robot
from orders.models import Order
from orders.signals import need_robot_signal
from .signals import robot_created_signal


@method_decorator(csrf_exempt, name='dispatch')
def create_robot(request):
    # Если метод запроса POST, то добавляем нового робота
    if request.method == 'POST':
        try:
            body = json.loads(request.body)  # Получаем данные из json

            # Достаём отдельные данные и записываем их в переменные
            robot_serial = body.get('serial')
            robot_model = body.get('model')
            robot_version = body.get('version')
            robot_created = body.get('created')
            robot_quantity = body.get('quantity')

            # Формируем словарь из полученных выше данных
            robot_data = {
                "serial": robot_serial,
                "model": robot_model,
                "version": robot_version,
                "created": robot_created,
                "quantity": robot_quantity
            }

            # Если робот, которого пользователь хочет создать есть в БД, достаём поле количество и добавляем в него
            # n-ое количество роботов
            if Robot.objects.filter(model=robot_model, version=robot_version).exists():
                old_quantity = Robot.objects.get(model=robot_model, version=robot_version)
                old_quantity.quantity += robot_quantity
                old_quantity.save()

                # И выводим сообщение о том, что пользователь добавил n-ое количество роботов на склад
                message = {
                    'message': f'Вы добавили на склад {robot_model}-{robot_version}, из серии: {robot_serial} в '
                               f'количестве {robot_quantity}. Теперь их всего: {old_quantity.quantity}'
                }
            else:
                # Если робот, которого пользователь хочет создать нет в БД, создаём его
                Robot.objects.create(**robot_data)

                # И выводим сообщение о том, что пользователь добавил робота в n-ом количестве
                message = {
                    'message': f'Вы создали нового робота: {robot_model}-{robot_version}, из серии: {robot_serial} '
                               f'в количестве {robot_quantity}.',
                }

        # Если пользователь ввел данные неверно выводим сообщение о том, что пользователь ошибся
        except:
            message = {
                'message': 'Ошибка при создании нового робота, проверьте все ли данные введены верно.',
            }

        # Возвращаем объект класса JsonResponse со статусом 201
        return JsonResponse(message, status=201)


# Если покупатель хотел заказать робота, но его не оказалось на складе, принимаем сигнал, о том, что нужно создать
# нового робота
@receiver(need_robot_signal)
def create_robot_for_order(sender, order_customer_email, order_serial, order_model, order_version, order_created,
                           order_quantity, **kwargs):
    # Создаём нужного робота в нужном количестве
    robot = Robot.objects.create(
        serial=order_serial,
        model=order_model,
        version=order_version,
        created=order_created,
        quantity=order_quantity
    )

    # Отправляем сигнал, о том, что робот создан в приложение "customers" и отправляем письмо на email покупателю
    robot_created_signal.send(sender=create_robot_for_order, order_customer_email=order_customer_email,
                              robot_model=order_model, robot_version=order_version)

    # Возвращаем созданного робота
    return robot


def download_excel_report(request):
    # Достаём из БД все модели роботов и делаем их множеством
    robots_models = set(Robot.objects.values_list('model', flat=True))

    # Создаём файл excel
    wb = Workbook()

    bold_font = Font(bold=True)  # Создаём объект класса Font, который будет отвечать за жирный шрифт в файле
    central_alignment = Alignment(horizontal='center',
                                  vertical='center')  # Создаём объект класса Alignment, который будет отвечать за
    # выравнивание текста по центру в файле

    # Создаём объект класса Border, который будет отвечать за обрисовывание клеток черным цветом
    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    # Проходимся циклом по множеству моделей роботов
    for robot_model in robots_models:

        robot_data = Robot.objects.values().filter(model=robot_model)  # Находим информацию о роботе с помощью модели

        ws = wb.create_sheet(robot_model)  # Создаём рабочий лист и в качестве названия берем модель
        # Увеличиваем длину столбцов
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 25

        # Заполняем первую строку названиями столбцов
        ws['A1'] = 'Модель'
        ws['B1'] = 'Версия'
        ws['C1'] = 'Количество за неделю'

        # Проходимся циклом по информации о роботе (перед этим превращаем queryset в список)
        for robot in list(robot_data):
            robot_model = robot['model']  # В переменную записываем модель робота
            robot_version = robot['version']  # В переменную записываем версию робота
            # В конец рабочего листа добавляем модель, версию робота и количество за неделю
            ws.append([robot_model, robot_version,
                       Order.objects.filter(
                           robot_model=robot_model,
                           robot_version=robot_version,
                           order_created__week=str(datetime.today().isocalendar()[1])
                       ).aggregate(Sum('robot_quantity'))['robot_quantity__sum']  # С помощью класса Sum складываем
                       # сумму количества заказов определенного робота за неделю
                       ])

        # Проходимся по ячейкам: A1, B1, C1 и делаем в них шрифт жирным
        for row in ws['A1':'C1']:
            for cell in row:
                cell.font = bold_font

        # Проходимся по всем строкам, начиная со второй, и в каждой ячейке делаем выравнивание текста по центру
        # и обрисовываем границы ячеек
        for row in ws.iter_rows(min_row=1, max_row=robot_data.count() + 1, max_col=3):
            for cell in row:
                cell.border = thin_border
                cell.alignment = central_alignment

    # При создании файла создаётся первый пустой рабочий лист, мы его удаляем
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])

    # Создаём имя файла, формируем его из даты и времени, когда он был скачан
    filename = f'all_robot_report {datetime.now().month}-{datetime.now().day}-{datetime.now().year} ' \
               f'{datetime.now().hour}_{datetime.now().minute}_{datetime.now().second}.xlsx'
    # Сохраняем файл и отдаём пользователю
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = f'attachment; filename={filename}'
    wb.save(response)
    return response
